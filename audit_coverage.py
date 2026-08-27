#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
生成"审计覆盖表"：对指定证书逐条给出 zlint 规则的处置结论。

审计完整性的定义: 规则清单里每一条规则都要有明确结论——
  - 已执行: pass / error / warn / NA / NE / info (直接来自 zlint 输出)
  - 未执行: 标注原因
      * "不适用(guard未过)" —— zlint 有该规则, 但证书不满足前置条件(CheckApplies), 即使 -includeNames 强制也不输出
      * "版本无此规则"     —— 该 lint 名不在当前 zlint 版本中 (规则清单按其他版本导出)
  - 额外一行: zlint 实际有、但规则清单未收录的规则 (标注"清单未收录")

用法:
    # 方式 A: 一个矩阵, 每张证书一列
    python audit_coverage.py --zlint zlint.exe --cert cert1.pem cert2.pem --out coverage.xlsx
    python audit_coverage.py --zlint zlint.exe --dir 文件夹 [--pattern "*.pem"] --out coverage.xlsx

    # 方式 B: 批量, 每张证书单独生成一个覆盖表到指定文件夹
    python audit_coverage.py --zlint zlint.exe --dir 文件夹 --out-dir 输出文件夹

输出: 矩阵(行=规则, 列=证书) 或 每证书一个 <证书名>.audit.xlsx; 均含汇总行。
"""
import argparse
import glob
import json
import os
import subprocess
import sys

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

CERT_EXTS = (".pem", ".der", ".crt", ".cer")

# 结果 -> 单元格底色 (与 run_zlint.py 一致)
RESULT_FILL = {
    "error": "FF9999",
    "fatal": "FF9999",
    "warn": "FFEB9C",
    "pass": "C6EFCE",
    "NA": "D9D9D9",
    "NE": "D9D9D9",
    "info": "DDEBF7",
    "不适用(guard未过)": "F2F2F2",
    "版本无此规则": "FFF2CC",
    "清单未收录": "E2EFDA",
}

RESULT_ZH = {
    "error": "错误", "fatal": "致命", "warn": "警告", "pass": "通过",
    "NA": "不适用", "NE": "未实现", "info": "信息",
}


def run_capture(cmd):
    proc = subprocess.run(cmd, capture_output=True)
    return proc.stdout.decode("utf-8", errors="replace"), proc.stderr.decode("utf-8", errors="replace")


def get_zlint_lints(zlint_bin):
    """返回 zlint 实际注册的 lint: {name: {"description":.., "source":..}}"""
    out, _ = run_capture([zlint_bin, "-list-lints-json"])
    lints = {}
    for line in out.splitlines():
        line = line.strip()
        if not line:
            continue
        try:
            obj = json.loads(line)
            lints[obj["name"]] = {"description": obj.get("description", ""),
                                  "source": obj.get("source", "")}
        except Exception:
            continue
    return lints


def run_zlint_on_cert(zlint_bin, cert):
    """跑默认 zlint, 返回 {lint名: result} (仅 JSON 部分)"""
    out, _ = run_capture([zlint_bin, cert])
    try:
        data = json.loads(out)
    except Exception:
        print(f"[警告] 无法解析 zlint 输出: {cert}", file=sys.stderr)
        return {}
    return {k: (v.get("result", "") if isinstance(v, dict) else "") for k, v in data.items()}


def find_certs(folder, pattern):
    files = []
    for root, _dirs, names in os.walk(folder):
        for n in names:
            if n.lower().endswith(CERT_EXTS):
                files.append(os.path.join(root, n))
    if pattern:
        import fnmatch
        files = [f for f in files if fnmatch.fnmatch(f.lower(), pattern.lower())]
    return sorted(set(os.path.abspath(f) for f in files))


def build_workbook(inventory, zlint_lints, rule_names, cert_results, certs):
    """构建覆盖表 workbook。行=规则, 列=证书; 末尾附逐证书统计。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "覆盖表"

    header = ["#", "Lint 名称", "规则描述", "来源"] + [os.path.basename(c) for c in certs]
    ws.append(header)

    for i, name in enumerate(rule_names, 1):
        row = []
        for cert in certs:
            res = cert_results[cert].get(name)
            if res is None:
                cell = "版本无此规则" if name not in zlint_lints else "不适用(guard未过)"
            else:
                cell = res if res else "?"
            row.append(cell)
        inv = inventory.get(name, {})
        zh = inv.get("zh", "") if isinstance(inv, dict) else ""
        src = inv.get("source", "") if isinstance(inv, dict) else ""
        if name in inventory:
            src_col = "规则清单" + (f" / {src}" if src else "")
        else:
            src_col = "zlint实际有(清单未收录)"
        ws.append([i, name, zh or zlint_lints.get(name, {}).get("description", ""), src_col] + row)

    # 逐证书统计 (行: 统计项, 列: 证书)
    n_version_missing = sum(1 for nm in rule_names if nm not in zlint_lints)
    stats = ["已执行(pass/error/warn/NA/NE/info)", "通过(pass)", "问题(error/fatal)", "警告(warn)",
             "不适用(guard未过)", "版本无此规则", "总计规则"]
    ws.append([])
    ws.append(["统计项"] + [os.path.basename(c) for c in certs])
    for stat in stats:
        row = [stat]
        for cert in certs:
            cr = cert_results[cert]
            n = len(cr)
            if stat.startswith("已执行"):
                row.append(n)
            elif stat == "通过(pass)":
                row.append(sum(1 for v in cr.values() if v == "pass"))
            elif stat == "问题(error/fatal)":
                row.append(sum(1 for v in cr.values() if v in ("error", "fatal")))
            elif stat == "警告(warn)":
                row.append(sum(1 for v in cr.values() if v == "warn"))
            elif stat == "不适用(guard未过)":
                row.append(len(rule_names) - n - n_version_missing)
            elif stat == "版本无此规则":
                row.append(n_version_missing)
            else:  # 总计规则
                row.append(len(rule_names))
        ws.append(row)

    # 样式
    from openpyxl.utils import get_column_letter
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 48
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 28
    for ci in range(5, 5 + len(certs)):
        ws.column_dimensions[get_column_letter(ci)].width = 14
    for r in ws.iter_rows(min_row=2, max_row=1 + len(rule_names)):
        for cell in r:
            if cell.column >= 5:
                fill = RESULT_FILL.get(cell.value)
                if fill:
                    cell.fill = PatternFill("solid", fgColor=fill)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "E2"
    return wb


def main():
    parser = argparse.ArgumentParser(description="生成审计覆盖表 (每张证书 × 每条 zlint 规则的处置结论)")
    parser.add_argument("--zlint", default="zlint.exe", help="zlint 可执行文件路径")
    parser.add_argument("--cert", nargs="+", help="一个或多个证书文件")
    parser.add_argument("--dir", help="证书文件夹 (递归)")
    parser.add_argument("--pattern", help="与 --dir 联用的文件匹配, 如 '*/positive/*.pem'")
    parser.add_argument("--out", default="audit_coverage.xlsx", help="方式A: 输出 xlsx 路径")
    parser.add_argument("--out-dir", default=None,
                        help="方式B: 输出文件夹, 每张证书单独生成一个 <证书名>.audit.xlsx")
    parser.add_argument("--rules", default=None, help="规则清单 json (默认 zlint-runner/lint_rules.json)")
    args = parser.parse_args()

    if not args.cert and not args.dir:
        print("[错误] 需要 --cert 或 --dir 指定证书。")
        sys.exit(1)

    zlint_bin = args.zlint
    if not os.path.isfile(zlint_bin):
        print(f"[错误] 找不到 zlint: {zlint_bin}")
        sys.exit(1)

    certs = []
    if args.dir:
        certs = find_certs(args.dir, args.pattern)
    else:
        certs = [os.path.abspath(c) for c in args.cert]
    certs = [c for c in certs if os.path.isfile(c)]
    if not certs:
        print("[错误] 没有可用的证书文件。")
        sys.exit(1)

    # 1) 规则清单 (兼容源码与打包 exe 两种场景)
    base_dir = os.path.dirname(os.path.abspath(__file__))
    candidates = [
        args.rules,
        os.path.join(base_dir, "lint_rules.json"),
        os.path.join(base_dir, "zlint-runner", "lint_rules.json"),
        os.path.join(base_dir, "..", "zlint-runner", "lint_rules.json"),
    ]
    rules_path = next((p for p in candidates if p and os.path.isfile(p)), None)
    inventory = {}
    if rules_path:
        with open(rules_path, encoding="utf-8") as f:
            inventory = json.load(f)
    else:
        print("[警告] 未找到规则清单 lint_rules.json (将只按 zlint 实际规则生成)")
    print(f"规则清单条数: {len(inventory)}")

    # 2) zlint 实际 lint
    zlint_lints = get_zlint_lints(zlint_bin)
    print(f"zlint {os.path.basename(zlint_bin)} 实际 lint 数: {len(zlint_lints)}")

    # 3) 规则全集 (并集, 保序: 先清单, 后 zlint 独有)
    rule_names = []
    seen = set()
    for n in inventory:
        rule_names.append(n)
        seen.add(n)
    for n in sorted(zlint_lints):
        if n not in seen:
            rule_names.append(n)
            seen.add(n)
    print(f"覆盖表规则总数 (并集): {len(rule_names)}")

    # 4) 逐证书跑
    print(f"证书数: {len(certs)}")
    cert_results = {}
    for cert in certs:
        print(f"  跑 {os.path.basename(cert)} ...")
        cert_results[cert] = run_zlint_on_cert(zlint_bin, cert)

    # 5) 输出
    if args.out_dir:
        try:
            os.makedirs(args.out_dir, exist_ok=True)
        except Exception as e:
            print(f"[错误] 无法创建输出目录 {args.out_dir}: {e}")
            sys.exit(1)
        saved = 0
        for cert in certs:
            wb = build_workbook(inventory, zlint_lints, rule_names,
                                {cert: cert_results[cert]}, [cert])
            stem = os.path.splitext(os.path.basename(cert))[0]
            path = os.path.join(args.out_dir, stem + ".audit.xlsx")
            try:
                wb.save(path)
                saved += 1
            except Exception as e:
                print(f"[警告] 保存失败 {path}: {e}")
        print(f"\n[已保存] {saved}/{len(certs)} 张覆盖表到 {args.out_dir}")
    else:
        wb = build_workbook(inventory, zlint_lints, rule_names, cert_results, certs)
        try:
            wb.save(args.out)
            print(f"\n[已保存] {args.out}")
            print(f"  - 规则总数: {len(rule_names)} 行 (清单 {len(inventory)} ∪ zlint {len(zlint_lints)})")
            print(f"  - 证书数: {len(certs)} 列")
            print("  - 单元格含义: pass/error/warn/NA/NE/info = zlint 实际结论;")
            print("    不适用(guard未过) = 该规则对证书不适用, zlint 不执行;")
            print("    版本无此规则 = 清单有此名但本版 zlint 没有;")
            print("    清单未收录 = zlint 有该规则但规则清单未收录。")
        except Exception as e:
            print(f"[错误] 保存 Excel 失败: {e}")
            sys.exit(1)

    try:
        input("\n按回车退出...")
    except EOFError:
        pass


if __name__ == "__main__":
    main()
