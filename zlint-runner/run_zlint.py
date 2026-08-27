#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
交互式运行 zlint 检查证书，并保存输出结果到文件。

运行时按提示操作:
    1) zlint 可执行文件的位置 (如 C:\...\zlint.exe 或 /c/...\zlint.exe)
    2) 是否批量检查一个文件夹里的所有证书? (y/N)
       - 选 y: 输入文件夹位置，递归遍历其中所有证书 (.pem/.der/.crt/.cer)
       - 选 n: 输入单个证书文件位置
    3) (可选) 只运行指定的 lint (-includeNames)，留空表示全部

结果会同时打印到屏幕，并保存为 <证书名>.zlint.json (默认)；
批量模式还会额外生成 <文件夹名>.batch.summary.json 汇总所有命中。

也支持纯命令行 (无交互):
    python run_zlint.py --dir 文件夹 [--zlint 路径] [--include lint] [--pattern "*.pem"]
    python run_zlint.py --cert 证书文件 [--zlint 路径] [--include lint]
"""

import argparse
import glob
import json
import os
import subprocess
import sys

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

OUTPUT_SUFFIX = ".zlint.json"
CERT_EXTS = (".pem", ".der", ".crt", ".cer")

# zlint 结果码 -> 中文含义，用于在 Excel 中"描述检测内容"
RESULT_MEANING = {
    "error": "错误：证书不符合规范要求，应被拒绝",
    "fatal": "致命错误：解析/结构层面无法处理",
    "warn": "警告：不推荐但不一定违规",
    "pass": "通过：符合该条检查要求",
    "info": "信息：仅供参考",
    "NA": "不适用：该检查对此证书不适用",
    "notice": "提示：需注意",
}
RESULT_FILL = {
    "error": "FFC7CE",   # 红
    "fatal": "FF9999",   # 深红
    "warn": "FFEB9C",    # 黄
    "pass": "C6EFCE",    # 绿
    "NA": "D9D9D9",      # 灰
}

# lint 名称 -> 中文含义（来源: zlint-detection-guide.md 详解）。
# 未列出的名称会走 guess_desc() 关键词兜底翻译。
LINT_DESC = {
    # 1. 基础字段
    "e_serial_number_not_positive": "序列号必须为正整数（不能为 0 或负数），且不超过 20 字节",
    "e_cert_extensions_version_not_3": "含扩展的证书版本必须是 v3",
    "e_cert_unique_identifier_version_not_2_or_3": "唯一标识符字段只在 v2/v3 中允许",
    "e_cert_contains_unique_identifier": "不允许包含 subjectUniqueID/issuerUniqueID 字段",
    "e_cert_ext_invalid_der": "扩展字段必须使用合法的 DER 编码",
    "e_basic_constr_invalid_der": "BasicConstraints 扩展的 DER 编码必须合法",
    "e_eku_critical_improperly": "部分扩展被错误地标记为 critical",
    "e_ext_duplicate_extension": "同一 OID 扩展不允许重复出现",
    "e_ext_cannot_be_empty_seq": "扩展值不允许是空序列",
    # 2. 有效期
    "e_sub_cert_valid_time_longer_than_398_days": "2020-09-01 后签发的 TLS 服务器证书有效期不得超过 398 天",
    "w_sub_cert_valid_time_longer_than_825_days": "早期证书有效期不得超过 825 天（历史要求）",
    "e_validity_time_not_positive": "证书有效期长度必须为正",
    # 3. Subject / Issuer
    "e_ca_subject_field_empty": "CA 证书的 Subject 不允许为空",
    "e_ca_country_name_missing": "CA 证书必须有合法的国家代码",
    "e_ca_country_name_invalid": "CA 证书的国家代码必须合法（2 字母 ISO）",
    "e_ca_common_name_missing": "根 CA 证书必须包含 Common Name",
    "e_ca_organization_name_missing": "CA 证书必须有组织名",
    "e_ca_dns_name_invalid": "Subject CN 中的 DNS 名称必须合法",
    "e_subject_contains_noninformational_value": "Subject 中不允许包含私人邮箱、电话等非机构信息",
    "w_subject_common_name_included": "CN 不建议保留（应迁移到 SAN）",
    "e_subject_common_name_not_from_san": "新规下 CN 应从 SAN 中复制，不允许独立设置",
    "e_subject_contains_reserved_arp": "不允许包含保留的 ARPA 域名",
    "e_subject_contains_reserved_ip": "不允许包含保留 IP 地址",
    "e_cab_ov_requires_org": "OV 证书必须包含组织名",
    "e_cab_iv_requires_personal_name": "IV（个人验证）证书必须有个人姓名",
    "e_ca_multiple_reserved_policy_oids": "不允许同时使用多个 CABF 保留策略 OID",
    # 4. DNS / SAN
    "e_dnsname_label_too_long": "域名标签长度不得超过 63 字符",
    "e_dnsname_contains_empty_label": "不允许空标签（如 foo..com）",
    "e_dnsname_contains_bare_iana_suffix": "不允许裸 TLD（如只签 com）",
    "e_dnsname_right_label_valid_tld": "最右侧标签必须是有效 TLD",
    "e_dnsname_hyphen_in_sld": "SLD 中不允许连字符开头/结尾",
    "e_dnsname_underscore_in_sld": "SLD 中不允许下划线",
    "e_dnsname_underscore_in_trd": "三级域名中不允许下划线",
    "e_dnsname_bad_character_in_label": "标签中不允许非法字符",
    "e_dnsname_contains_prohibited_reserved_label": "不允许保留标签",
    "e_dnsname_wildcard_left_of_public_suffix": "通配符必须覆盖整个公共后缀左侧（不允许 *.com）",
    "e_dnsname_check_left_label_wildcard": "通配符 *. 只允许出现在最左侧标签",
    "e_ext_san_dns_name_valid": "SAN 中的 DNS 名称必须可解析且符合规范",
    "e_arpa_domain_not_allowed": "不允许使用 .arpa 保留域名",
    # 5. 公钥 / 签名算法
    "e_rsa_mod_less_than_2048_bits": "RSA 密钥长度不得小于 2048 位",
    "e_rsa_public_exponent_too_small": "RSA 公钥指数 E 过小（应为 ≥ 65537 且为奇数）",
    "e_rsa_public_exponent_not_odd": "RSA 公钥指数 E 必须为奇数",
    "e_rsa_mod_not_odd": "RSA 模数必须是奇数",
    "e_ecdsa_allowed_ku": "ECDSA 证书的 KeyUsage 必须符合规范",
    "e_ecdsa_ee_invalid_ku": "ECDSA 终端实体证书的 KeyUsage 不合法",
    "e_signature_algorithm_not_supported": "签名算法必须是已知/受支持的",
    "w_rsa_public_exponent_too_small": "RSA 公钥指数过小的警告",
    "e_dh_params_missing": "DH 证书必须包含 DH 参数",
    "e_old_sub_cert_rsa_mod_less_than_1024_bits": "2014 年前签发证书 RSA 不得小于 1024 位",
    # 6. 扩展
    "e_basic_constraints_not_critical": "CA 证书的 BasicConstraints 必须标记为 critical",
    "w_basic_constraints_not_critical": "叶子证书若含 BasicConstraints 也应 critical",
    "e_is_ca": "根/中间 CA 必须 cA=TRUE",
    "e_ca_key_usage_missing": "CA 证书必须包含 KeyUsage",
    "e_ca_key_usage_not_critical": "CA 证书的 KeyUsage 必须 critical",
    "e_ca_key_cert_sign_not_set": "CA 证书必须设置 keyCertSign 位",
    "e_ca_crl_sign_not_set": "CA 证书必须设置 cRLSign 位",
    "e_ca_digital_signature_not_set": "CA 证书必须设置 digitalSignature 位",
    "e_sub_cert_key_usage_cert_sign": "叶子证书禁止设置 keyCertSign",
    "e_sub_cert_eku_server_auth_client_auth": "不允许同时包含 serverAuth 和 clientAuth",
    "e_subject_key_identifier_missing": "所有证书必须包含 SKI（主题密钥标识符）",
    "w_subject_key_identifier_missing": "CA 证书必须包含 SKI（早期要求）",
    "e_subject_key_identifier_not_20_bytes": "SKI 必须是 20 字节规范形式",
    "e_ext_authority_key_identifier_no_key_identifier": "AKI 必须包含 keyIdentifier",
    "e_ext_authority_key_identifier_critical": "AKI 不允许标记 critical",
    "e_ca_akid_key_identifier_missing": "CA 证书 AKI 必须与签发者 SKI 匹配",
    "e_aia_ocsp_must_have_http_only": "AIA 中 OCSP URL 必须仅为 HTTP",
    "e_aia_ca_issuers_must_have_http_only": "AIA 中 caIssuers URL 必须仅为 HTTP",
    "e_aia_must_contain_permitted_access_method": "AIA 必须包含允许的访问方法",
    "e_aia_unique_locations": "AIA 中不允许重复位置",
    "e_ext_aia_access_location_missing": "AIA 条目必须包含 location",
    "e_ext_aia_marked_critical": "AIA 不允许 critical",
    "e_crl_distrib_points_not_http": "CDP（CRL 分发点）必须为 HTTP",
    "e_crl_distrib_points_marked_critical": "CDP 不允许 critical",
    "e_crlissuer_must_not_be_present_in_cdp": "CDP 中不允许包含 cRLIssuer",
    "w_distribution_point_incomplete": "DP 必须完整（name+reasons+cRLIssuer 一致）",
    "e_ext_cert_policy_duplicate": "证书策略不允许重复",
    "e_ext_cert_policy_disallowed_any_policy_qualifier": "anyPolicy 不允许带 qualifier",
    "e_ext_cert_policy_contains_noticeref": "不允许使用 noticeRef（仅允许 CPS/explicitText）",
    "e_empty_sct_list": "SCT 列表不允许为空",
    "e_qcstatem_qccompliance_valid": "ETSI QcCompliance 声明必须格式正确",
    "e_qcstatem_qclimitvalue_valid": "QC 限制值（金额限制）声明必须正确编码",
    "e_qcstatem_qcretentionperiod_valid": "证书保留期声明必须合法",
    "e_qcstatem_qcsscd_valid": "私钥存储在 SSCD 的声明必须合法",
    # 7. CRL
    "e_crl_missing_crl_number": "CRL 必须包含 CRL Number 扩展",
    "e_crl_has_authority_key_identifier": "CRL 必须包含 AKI 扩展",
    "e_crl_has_next_update": "CRL 必须包含 nextUpdate 字段",
    "e_crl_empty_revoked_certificates": "CRL 中 revokedCertificates 不允许为空",
    "e_crl_revoked_certificates_field_empty": "CRL 中 revokedCertificates 字段不允许为空",
    "e_crl_revocation_time_not_after_this_update": "吊销时间不能晚于 thisUpdate",
    "e_crl_valid_reason_codes": "吊销原因码必须合法（0-10）",
    "e_cabf_crl_valid_reason_codes": "CABF 下吊销原因码必须合法",
    "e_crl_sigalgo_missing_null_params": "CRL 签名算法的 RSA 参数必须包含 NULL",
    "e_crl_number_range": "CRL Number 必须为非负整数且在 20 字节内",
    "e_crl_next_update_invalid": "nextUpdate 必须晚于 thisUpdate",
    "e_crl_extensions": "CRL 扩展必须合法（CABF BR 规定只能含指定扩展）",
    "e_crl_auth_key_id_only_contains_keyid": "CRL 的 AKI 只能包含 keyIdentifier",
    "e_cabf_crl_reason_code_not_critical": "吊销原因扩展不应 critical",
    "e_crl_this_update_in_future": "thisUpdate 不能在未来",
    # 10. Root Store
    "e_tls_server_cert_valid_time_longer_than_398_days": "TLS 服务器证书有效期 ≤ 398 天",
    "e_server_auth_eku": "服务器证书必须含 serverAuth EKU",
}

# 关键词兜底翻译（当 LINT_DESC 未精确命中时使用）
_KEYWORDS = [
    ("serial", "序列号"), ("version", "版本"), ("validity", "有效期"), ("utc", "UTC时间"),
    ("generalized", "GeneralizedTime"), ("time", "时间"), ("subject", "主体(Subject)"),
    ("issuer", "签发者(Issuer)"), ("common_name", "通用名称(CN)"), ("country", "国家"),
    ("organization", "组织"), ("dn", "可分辨名称"), ("dns", "DNS名称"), ("san", "主题备用名(SAN)"),
    ("ian", "签发者备用名(IAN)"), ("rsa", "RSA"), ("ecdsa", "ECDSA"), ("ec", "椭圆曲线"),
    ("dsa", "DSA"), ("dh", "DH"), ("key", "密钥"), ("signature", "签名"), ("algorithm", "算法"),
    ("mod", "模数"), ("exponent", "指数"), ("curve", "曲线"), ("ext", "扩展"), ("extension", "扩展"),
    ("critical", "critical 标记"), ("policy", "策略"), ("policies", "策略"), ("ocsp", "OCSP"),
    ("aia", "权威信息访问(AIA)"), ("crl", "证书吊销列表(CRL)"), ("sct", "证书透明(SCT)"),
    ("qc", "合格证书(QC)"), ("skid", "主题密钥标识符(SKI)"), ("aki", "签发者密钥标识符(AKI)"),
    ("cabf", "CA/B Forum"), ("mozilla", "Mozilla"), ("apple", "Apple"), ("chrome", "Chrome"),
    ("ev", "扩展验证(EV)"), ("smime", "S/MIME"), ("code_sign", "代码签名"), ("ev_", "EV"),
    ("ca", "证书颁发机构(CA)"), ("sub", "订阅/叶子"), ("ee", "终端实体"), ("root", "根证书"),
    ("empty", "为空"), ("missing", "缺失"), ("invalid", "不合法"), ("duplicate", "重复"),
    ("not_positive", "非正"), ("less_than", "过小"), ("longer_than", "过长"), ("must", "必须"),
    ("not", "不"), ("contains", "包含"), ("allowed", "允许"), ("disallowed", "不允许"),
    ("wildcard", "通配符"), ("label", "标签"), ("reserved", "保留"), ("bare", "裸"),
    ("hyphen", "连字符"), ("underscore", "下划线"), ("bad_character", "非法字符"),
    ("name", "名称"), ("cert", "证书"), ("number", "编号"), ("reason", "原因码"),
    ("explicit_text", "显式文本"), ("qualifier", "限定符"), ("notice", "通知"),
    ("length", "长度"), ("bytes", "字节"), ("range", "范围"), ("future", "未来"),
    ("exponent_too_small", "指数过小"),
]


def guess_desc(name):
    """对未在 LINT_DESC 中列出的 lint 名，用关键词做近似中文释义。"""
    low = name.lower()
    parts = [zh for en, zh in _KEYWORDS if en in low]
    if parts:
        prefix = "错误检查: " if name.startswith("e_") else (
            "警告检查: " if name.startswith("w_") else "")
        return prefix + "检查" + "".join(parts)
    return ""


def _load_rules_json():
    """从脚本同目录的 lint_rules.json 加载 432 条 lint 规则描述。
    该 JSON 由 zlint-lint规则详解.xlsx 导出，找不到时返回 {}。
    """
    try:
        base = os.path.dirname(os.path.abspath(__file__))
        path = os.path.join(base, "lint_rules.json")
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:  # noqa: BLE001
        return {}


_RULES = _load_rules_json()


def lint_desc(name):
    # 1) 优先: lint_rules.json (由 zlint-lint规则详解.xlsx 导出，432 条)
    r = _RULES.get(name)
    if r:
        zh = r.get("zh") if isinstance(r, dict) else r
        if zh:
            return zh
    # 2) 其次: 内置字典 (zlint-detection-guide.md)
    # 3) 最后: 关键词兜底
    return LINT_DESC.get(name) or guess_desc(name)


def prompt_path(label, must_exist=True):
    """让用户输入一个路径，做基本的存在性/可读性检查。"""
    while True:
        val = input(f"{label}: ").strip().lstrip("\ufeff").strip('"').strip("'")
        if not val:
            print("  请输入一个有效路径，不能为空。")
            continue
        if must_exist and not os.path.exists(val):
            print(f"  路径不存在: {val}")
            retry = input("  仍要继续? (y/N): ").strip().lower()
            if retry != "y":
                continue
            return val
        return val


def write_xlsx(base, data, cert_name=""):
    """把 zlint 的 JSON 结果写成 Excel 表格，描述每条检测内容。

    列: 序号 | Lint 名称 | 结果 | 结果含义 | 详情
    返回生成的 xlsx 路径；data 不是 dict 时返回 None。
    """
    if not isinstance(data, dict):
        return None

    rows = []
    for name, info in data.items():
        if not isinstance(info, dict):
            continue
        result = info.get("result", "")
        details = info.get("details", "") or ""
        rows.append((name, result, RESULT_MEANING.get(result, result), details, lint_desc(name)))

    wb = Workbook()
    ws = wb.active
    ws.title = "检测结果"
    headers = ["序号", "Lint 名称", "结果", "结果含义", "检测内容说明", "详情"]
    ws.append(headers)
    hdr_fill = PatternFill("solid", fgColor="305496")
    hdr_font = Font(bold=True, color="FFFFFF")
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(vertical="center")

    for i, (name, result, meaning, details, desc) in enumerate(rows, 1):
        ws.append([i, name, result, meaning, desc, details])
        fill = RESULT_FILL.get(result)
        if fill:
            ws.cell(row=i + 1, column=3).fill = PatternFill("solid", fgColor=fill)

    # 列宽
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 48
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 60
    ws.column_dimensions["F"].width = 60
    ws.freeze_panes = "A2"

    xlsx_path = base + ".zlint.xlsx"
    try:
        wb.save(xlsx_path)
    except Exception as e:  # noqa: BLE001
        print(f"[警告] 保存 Excel 失败: {e}")
        return None
    return xlsx_path


def run_one(zlint_bin, cert_path, include, verbose=True):
    """运行一次 zlint 并检查单个证书，保存结果；返回 (saved_json_path, flagged_dict)。"""
    cmd = [zlint_bin, cert_path]
    if include:
        cmd.insert(1, "-includeNames=" + include)

    if verbose:
        print("\n运行命令: " + " ".join(cmd))
        print("-" * 60)

    try:
        proc = subprocess.run(cmd, capture_output=True, text=True)
    except FileNotFoundError:
        print(f"[错误] 找不到可执行文件: {zlint_bin}")
        print("        请确认路径是否正确，或在 PATH 中可直接调用 zlint。")
        return None, None
    except Exception as e:  # noqa: BLE001
        print(f"[错误] 运行失败: {e}")
        return None, None

    stdout = proc.stdout
    stderr = proc.stderr

    if verbose:
        if stdout:
            print(stdout)
        if stderr:
            print("[stderr]\n" + stderr)
        if proc.returncode != 0:
            print(f"[提示] 进程返回码: {proc.returncode} (非 0 不一定代表失败，"
                  "zlint 对 error 结果常返回非 0)")

    base, _ = os.path.splitext(cert_path)
    out_path = base + OUTPUT_SUFFIX
    try:
        with open(out_path, "w", encoding="utf-8") as f:
            f.write(stdout)
            if stderr:
                f.write("\n\n[stderr]\n" + stderr)
        if verbose:
            print("-" * 60)
            print(f"[已保存] {out_path}")
    except Exception as e:  # noqa: BLE001
        print(f"[警告] 保存失败: {e}")

    flagged = {}
    try:
        data = json.loads(stdout)
        flagged = {
            k: v for k, v in data.items()
            if isinstance(v, dict) and v.get("result") in ("error", "warn", "fatal")
        }
        if flagged:
            summary_path = base + ".zlint.summary.json"
            with open(summary_path, "w", encoding="utf-8") as f:
                json.dump(flagged, f, indent=2, ensure_ascii=False)
            if verbose:
                print(f"[已保存] {summary_path} (仅含 error/warn/fatal)")
                print(f"         命中 {len(flagged)} 条检查")

        # 额外生成 Excel 表格，描述每条检测内容
        xlsx_path = write_xlsx(base, data)
        if xlsx_path and verbose:
            print(f"[已保存] {xlsx_path} (Excel 表格: 每条 lint 的结果与说明)")
    except json.JSONDecodeError:
        pass

    return out_path, flagged


def find_cert_files(folder, pattern=None):
    """返回文件夹（递归）下所有证书文件。pattern 为相对文件夹的 glob。"""
    files = []
    if pattern:
        files = glob.glob(os.path.join(folder, pattern), recursive=True)
    else:
        for root, _dirs, names in os.walk(folder):
            for n in names:
                if n.lower().endswith(CERT_EXTS):
                    files.append(os.path.join(root, n))
    files = [
        f for f in files
        if os.path.isfile(f)
        and not f.endswith(OUTPUT_SUFFIX)
        and not f.endswith(".zlint.summary.json")
    ]
    return sorted(set(os.path.abspath(f) for f in files))


def batch_mode(zlint_bin, folder, pattern, include):
    certs = find_cert_files(folder, pattern)
    if not certs:
        print(f"[错误] 在 {folder} 下未发现证书文件"
              + (f" (pattern={pattern})" if pattern else ""))
        sys.exit(1)

    print("=" * 60)
    print(f"  批量模式: 共发现 {len(certs)} 个证书")
    print(f"  文件夹: {folder}")
    if pattern:
        print(f"  匹配:  {pattern}")
    print("=" * 60)

    summary = {}
    for i, cert in enumerate(certs, 1):
        print(f"\n[{i}/{len(certs)}] {cert}")
        _out, flagged = run_one(zlint_bin, cert, include, verbose=False)
        if flagged:
            summary[cert] = flagged

    base = os.path.basename(os.path.normpath(folder))
    batch_summary_path = os.path.join(folder, f"{base}.batch.summary.json")
    with open(batch_summary_path, "w", encoding="utf-8") as f:
        json.dump(summary, f, indent=2, ensure_ascii=False)

    print("\n" + "=" * 60)
    print(f"完成。{len(certs)} 个证书已处理，"
          f"其中 {len(summary)} 个有 error/warn/fatal。")
    print(f"批量汇总: {batch_summary_path}")
    total = sum(len(v) for v in summary.values())
    print(f"命中检查总数: {total}")

    # 额外生成批量汇总 Excel：每行一个证书，列出命中数与命中的 lint
    batch_xlsx = os.path.join(folder, f"{base}.batch.xlsx")
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "批量汇总"
        ws.append(["证书", "命中数", "命中的 lint (error/warn/fatal)"])
        for c in range(1, 4):
            cell = ws.cell(row=1, column=c)
            cell.fill = PatternFill("solid", fgColor="305496")
            cell.font = Font(bold=True, color="FFFFFF")
        for cert, fl in summary.items():
            lint_names = ", ".join(
                f"{k}={v.get('result')}" for k, v in fl.items()
            )
            ws.append([os.path.basename(cert), len(fl), lint_names])
        ws.column_dimensions["A"].width = 50
        ws.column_dimensions["B"].width = 10
        ws.column_dimensions["C"].width = 80
        ws.freeze_panes = "A2"
        wb.save(batch_xlsx)
        print(f"批量汇总 Excel: {batch_xlsx}")
    except Exception as e:  # noqa: BLE001
        print(f"[警告] 批量汇总 Excel 生成失败: {e}")


def interactive_mode():
    print("=" * 60)
    print("  zlint 交互式证书检查工具")
    print("=" * 60)

    zlint_bin = prompt_path("请输入 zlint 可执行文件位置 (如 zlint.exe)")

    choice = input("是否批量检查一个文件夹里的所有证书? (y/N): ").strip().lstrip("\ufeff").lower()
    include = input("可选: 只运行某个 lint 名 (-includeNames，留空=全部): ").strip().lstrip("\ufeff")

    if choice == "y":
        folder = prompt_path("请输入证书文件夹位置 (将递归遍历所有证书)", must_exist=True)
        batch_mode(zlint_bin, folder, None, include)
    else:
        cert_path = prompt_path("请输入要检查的证书文件位置 (如 xxx.pem)")
        run_one(zlint_bin, cert_path, include, verbose=True)


def main():
    parser = argparse.ArgumentParser(
        description="运行 zlint 检查证书（交互式选择单张/批量，或命令行指定）")
    parser.add_argument("--dir", help="证书文件夹，批量模式（递归遍历）")
    parser.add_argument("--cert", help="单个证书文件，单张模式")
    parser.add_argument("--zlint", help="zlint 可执行文件路径")
    parser.add_argument("--include", help="只运行某个 lint 名 (-includeNames)")
    parser.add_argument("--pattern", help="批量模式的文件匹配 glob，如 '*/positive/*.pem'")
    args = parser.parse_args()

    if args.dir:
        if not args.zlint:
            print("[错误] --dir 模式需要 --zlint 指定 zlint 可执行文件。")
            sys.exit(1)
        if not os.path.isdir(args.dir):
            print(f"[错误] 文件夹不存在: {args.dir}")
            sys.exit(1)
        batch_mode(args.zlint, args.dir, args.pattern, args.include)
    elif args.cert:
        if not args.zlint:
            print("[错误] --cert 模式需要 --zlint 指定 zlint 可执行文件。")
            sys.exit(1)
        if not os.path.isfile(args.cert):
            print(f"[错误] 文件不存在: {args.cert}")
            sys.exit(1)
        run_one(args.zlint, args.cert, args.include, verbose=True)
    else:
        interactive_mode()

    try:
        input("\n按回车退出...")
    except EOFError:
        pass


if __name__ == "__main__":
    main()
